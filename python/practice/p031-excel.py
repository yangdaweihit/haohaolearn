#! /usr/bin/python3
# -*- coding: utf-8 -*-
"""
本脚本用于实验Python读写Excel表格


"""

# 使用Python进行Excel数据读写
# 示例文件：./files/data.xlsx

a = tuple()


def foo(x):
    return x, x**2


import pandas as pd
import xlrd
import xlsxwriter

import collections

Sale = collections.namedtuple("Sale",
                              "productid customeid date quantity price")


def foo(n):
    return list(range(n+1))

# 用pandas读入
excelfile = r"files/data.xlsx"
data = pd.read_excel(io=excelfile, sheet_name='Sheet1')
print(data)

# 打开一个表格
xl = xlrd.open_workbook(excelfile)
# 获取第1页
table = xl.sheets()[0]
print(table)
# 获取该页行数
rows = table.nrows
print(rows)
# 获取第1行
row = table.row_values(0)
print(row)
# 获取第1列
col = table.col_values(0)
print(col)
# 获取第2行第2列值
data = table.cell(1, 1).value
print(data)

from xlrd import open_workbook, cellname
book = open_workbook(excelfile)
sheet = book.sheet_by_index(0)
print(sheet.name)
print(sheet.nrows)
print(sheet.ncols)
for row_index in range(sheet.nrows):
    for col_index in range(sheet.ncols):
        print(cellname(row_index, col_index), '-')
        print(sheet.cell(row_index, col_index).value)
""" openpyxl令人最意外的功能是：你在操作一个表格时，甚至不必先生成一个文
件。

"""
from openpyxl import Workbook, load_workbook

wb = Workbook()
ws = wb.active
# Data can be assigned directly to cells
ws['A1'] = 42
# Rows can also be appended
# 增加一行
ws.append([1, 2, 3])
# 命名当前页
ws.title = "New Title"
ws.sheet_properties.tabColor = "1072BA"
# insert at the end (default)
ws1 = wb.create_sheet("Mysheet")
ws2 = wb["New Title"]
print(ws2)
print(wb.sheetnames)
for sheet in wb:
    print(sheet.title)

# This operation will overwrite existing files without warning.
# 这个操作将在没有警告的情况下覆盖当前文件。
wb.save('balances.xlsx')

# Loading from a file
# 加载一个文件
wb2 = load_workbook(excelfile)
print(wb2.sheetnames)

source = wb.active
target = wb.copy_worksheet(source)
print(target)
